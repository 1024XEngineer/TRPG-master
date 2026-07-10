from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
REPO_DIR = ROOT_DIR.parent
WORKBOOK_PATH = REPO_DIR / "COC7空白卡CY23Final.xlsx"
OUTPUT_PATH = ROOT_DIR / "src/data/generated/coc7-occupations.json"

OCCUPATION_SHEET = "职业列表"
OCCUPATION_SKILL_SHEET = "本职技能"

ATTRIBUTES = {
    "力量": "str",
    "体质": "con",
    "意志": "pow",
    "敏捷": "dex",
    "外貌": "app",
    "体型": "siz",
    "智力": "int",
    "教育": "edu",
    "幸运": "luck",
}

CATEGORY_RULES = [
    {"label": "医疗保健", "icon": "🏥", "keywords": ["医生", "护士", "医学", "药剂", "心理", "治疗", "救援", "牙医", "外科", "整形", "精神病"]},
    {"label": "执法安全", "icon": "🔒", "keywords": ["警", "探员", "侦探", "士兵", "军", "消防", "保安", "密探", "自卫队", "海警", "佣兵", "军官", "治安官"]},
    {"label": "法律金融", "icon": "⚖️", "keywords": ["会计", "律师", "法官", "司法", "证券", "银行", "保释", "担保", "政府官员"]},
    {"label": "文化艺术", "icon": "🎨", "keywords": ["艺术", "作家", "音乐", "摄影", "演员", "艺人", "舞者", "歌手", "设计师", "偶像", "播音", "主持", "评论家", "专栏", "撰稿", "舞台"]},
    {"label": "学术研究", "icon": "📚", "keywords": ["教授", "学者", "考古", "研究", "科学", "图书馆", "博物馆", "文物", "实验", "教师", "学生", "炼丹", "学识"]},
    {"label": "野外生存", "icon": "🏔️", "keywords": ["猎人", "探险", "农民", "渔民", "登山", "牛仔", "海员", "船员", "潜水", "山岳", "旅行", "飞行员", "司机", "车夫", "勘测", "寻宝"]},
    {"label": "社会边缘", "icon": "🎭", "keywords": ["罪犯", "黑帮", "走私", "混混", "非法", "暴走", "骇客", "黑客", "密医", "性工作", "自宅警备"]},
    {"label": "专业人员", "icon": "🔧", "keywords": ["工程师", "程序员", "技师", "机械师", "工匠", "建筑师", "工人", "劳工", "电话", "秘书", "店主", "店老板", "小企业家", "厨师", "管家", "仆", "佣人", "服务生", "白领"]},
    {"label": "社交服务", "icon": "🤝", "keywords": ["神职", "牧师", "记者", "酒保", "推销", "传教", "大使", "发言人", "顾问", "咨询", "工会", "狂热者", "占卜", "灵媒", "风水", "市子"]},
]

DEFAULT_CATEGORY = {"label": "其他职业", "icon": "🗂️", "keywords": []}

ICON_RULES = [
    ("会计", "📊"),
    ("考古", "🏛️"),
    ("文物", "🏺"),
    ("古董", "🏺"),
    ("艺术", "🎨"),
    ("演员", "🎭"),
    ("电影", "🎬"),
    ("作家", "✒️"),
    ("记者", "📰"),
    ("摄影", "📷"),
    ("音乐", "🎼"),
    ("歌手", "🎙️"),
    ("舞者", "💃"),
    ("医生", "⚕️"),
    ("护士", "💉"),
    ("药剂", "💊"),
    ("心理", "🧠"),
    ("律师", "⚖️"),
    ("法官", "⚖️"),
    ("警", "👮"),
    ("侦探", "🔎"),
    ("探员", "🕵️"),
    ("密探", "🕵️"),
    ("士兵", "🎖️"),
    ("军", "🎖️"),
    ("消防", "🚒"),
    ("教授", "🎓"),
    ("学者", "🎓"),
    ("学生", "🎒"),
    ("科学", "🔬"),
    ("图书馆", "📚"),
    ("博物馆", "🏛️"),
    ("神职", "⛪"),
    ("牧师", "⛪"),
    ("猎人", "🏹"),
    ("探险", "🧭"),
    ("旅行", "🧭"),
    ("登山", "⛰️"),
    ("飞行", "✈️"),
    ("司机", "🚗"),
    ("车夫", "🐎"),
    ("海员", "⚓"),
    ("船员", "⚓"),
    ("潜水", "🤿"),
    ("渔民", "🎣"),
    ("农民", "🌾"),
    ("运动员", "🏅"),
    ("拳击", "🥊"),
    ("摔跤", "🥊"),
    ("黑客", "💻"),
    ("骇客", "💻"),
    ("程序员", "💻"),
    ("工程师", "⚙️"),
    ("技师", "🔧"),
    ("机械", "🔧"),
    ("建筑师", "📐"),
    ("工匠", "🛠️"),
    ("厨师", "🍳"),
    ("管家", "🛎️"),
    ("仆", "🛎️"),
    ("酒保", "🍸"),
    ("推销", "🤝"),
    ("大使", "🏛️"),
    ("发言人", "🎤"),
    ("罪犯", "🎭"),
    ("黑帮", "🎭"),
    ("佣兵", "🎖️"),
    ("店", "🏪"),
]

SKILL_LABEL_TO_IDS = {
    "会计": ["accounting"],
    "人类学": ["anthropology"],
    "估价": ["appraise"],
    "考古学": ["archaeology"],
    "技艺①": ["art-craft-1"],
    "技艺②": ["art-craft-2"],
    "技艺③": ["art-craft-3"],
    "取悦": ["charm"],
    "攀爬": ["climb"],
    "计算机使用 Ω": ["computer-use"],
    "克苏鲁神话": ["cthulhu-mythos"],
    "乔装": ["disguise"],
    "闪避": ["dodge"],
    "汽车驾驶": ["drive-auto"],
    "电气维修": ["electrical-repair"],
    "电子学 Ω": ["electronics"],
    "话术": ["fast-talk"],
    "格斗：": ["fighting-brawl"],
    "格斗①": ["fighting-brawl"],
    "格斗②": ["fighting-sword"],
    "格斗③": ["fighting-knife"],
    "射击：": ["firearm-handgun"],
    "射击①": ["firearm-handgun"],
    "射击②": ["firearm-rifle"],
    "射击③": ["firearm-smg"],
    "急救": ["first-aid"],
    "历史": ["history"],
    "恐吓": ["intimidate"],
    "跳跃": ["jump"],
    "外语①": ["language-foreign-1"],
    "外语②": ["language-foreign-2"],
    "外语③": ["language-foreign-3"],
    "母语": ["language-native"],
    "法律": ["law"],
    "图书馆使用": ["library-use"],
    "聆听": ["listen"],
    "锁匠": ["lock-smith"],
    "机械维修": ["mechanical-repair"],
    "医学": ["medicine"],
    "博物学": ["natural-world"],
    "导航": ["navigation"],
    "神秘学": ["occult"],
    "操作重型机械": ["heavy-machinery"],
    "说服": ["persuade"],
    "驾驶：": ["pilot-aircraft"],
    "心理学": ["psychology"],
    "骑术": ["ride"],
    "科学①": ["science-biology"],
    "科学②": ["science-chemistry"],
    "科学③": ["science-physics"],
    "侦查": ["spot-hidden"],
    "潜行": ["stealth"],
    "生存：": ["survival"],
    "游泳": ["swim"],
    "投掷": ["throw"],
    "追踪": ["track"],
}


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"[ \t]+", " ", text)


def is_source_note(text: str | None) -> bool:
    if not text:
        return False
    return bool(re.search(r"使用前请征得KP同意|呼唤调查员伴侣|克苏鲁2015|克苏鲁煤气灯", text))


def trim_sentence(text: str, max_length: int) -> str:
    normalized = text.strip()
    if len(normalized) <= max_length:
        return normalized
    return normalized[: max_length - 1] + "…"


def classify_occupation(name: str) -> dict[str, Any]:
    for rule in CATEGORY_RULES:
        if any(keyword in name for keyword in rule["keywords"]):
            return rule
    return DEFAULT_CATEGORY


def icon_for_occupation(name: str, category: dict[str, Any]) -> str:
    for keyword, icon in ICON_RULES:
        if keyword in name:
            return icon
    return category["icon"]


def make_generated_description(name: str, category: str, skills_text: str | None) -> str:
    if skills_text:
        skill_summary = skills_text.replace("。", "").strip()
        return f"{name}属于{category}类职业，职业技能侧重于{skill_summary}。"
    return f"{name}属于{category}类职业，可根据守秘人的规则要求补充职业背景。"


def make_display_description(name: str, category: str, description: str | None, skills_text: str | None) -> str:
    if description and not is_source_note(description):
        return description
    return make_generated_description(name, category, skills_text)


def make_short_desc(description: str | None, skills_text: str | None) -> str:
    source = description or skills_text or "COC7 职业模板"
    first_line = source.splitlines()[0].strip()
    return trim_sentence(first_line, 46)


def parse_credit_range(value: str | None) -> dict[str, int] | None:
    if value is None:
        return None
    match = re.fullmatch(r"(\d+)\s*-\s*(\d+)", value)
    if not match:
        return None
    return {"min": int(match.group(1)), "max": int(match.group(2))}


def parse_point_formula(formula: str | None) -> dict[str, Any]:
    if not formula:
        return {"kind": "unknown", "raw": ""}

    normalized = (
        formula.replace("＋", "+")
        .replace("×", "*")
        .replace(" ", "")
        .replace("EDU", "教育")
        .replace("STR", "力量")
        .replace("DEX", "敏捷")
        .replace("APP", "外貌")
        .replace("POW", "意志")
    )

    fixed = re.fullmatch(r"([一-龥]+)\*([0-9]+)", normalized)
    if fixed and fixed.group(1) in ATTRIBUTES:
        return {
            "kind": "fixed",
            "attr": ATTRIBUTES[fixed.group(1)],
            "multiplier": int(fixed.group(2)),
            "raw": formula,
        }

    edu_plus = re.fullmatch(r"教育\*2\+(.+)\*2", normalized)
    if edu_plus:
        raw_options = [part for part in edu_plus.group(1).split("或") if part]
        attrs = [ATTRIBUTES[part] for part in raw_options if part in ATTRIBUTES]
        if len(attrs) == 1:
            return {
                "kind": "sum",
                "terms": [
                    {"attr": "edu", "multiplier": 2},
                    {"attr": attrs[0], "multiplier": 2},
                ],
                "raw": formula,
            }
        if len(attrs) > 1:
            return {
                "kind": "choice",
                "base": {"attr": "edu", "multiplier": 2},
                "options": attrs,
                "optionMultiplier": 2,
                "raw": formula,
            }

    return {"kind": "unknown", "raw": formula}


def collect_skill_matrix(workbook: Any) -> dict[int, list[dict[str, Any]]]:
    sheet = workbook[OCCUPATION_SKILL_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    occupation_rules: dict[int, list[dict[str, Any]]] = {}

    header = rows[0]
    for col_index, excel_no in enumerate(header[2:], start=3):
        if isinstance(excel_no, (int, float)):
            occupation_rules[int(excel_no)] = []

    for row_index, row_values in enumerate(rows[7:], start=8):
        label = clean_text(row_values[0] if row_values else None)
        if not label or label not in SKILL_LABEL_TO_IDS:
            continue

        for col_index, marker_value in enumerate(row_values[2:], start=3):
            excel_no = header[col_index - 1] if col_index - 1 < len(header) else None
            marker = clean_text(marker_value)
            if not isinstance(excel_no, (int, float)) or marker is None:
                continue

            for skill_id in SKILL_LABEL_TO_IDS[label]:
                occupation_rules.setdefault(int(excel_no), []).append(
                    {
                        "kind": "matrix",
                        "skillId": skill_id,
                        "label": label,
                        "marker": marker,
                        "source": {"sheet": OCCUPATION_SKILL_SHEET, "row": row_index, "column": col_index},
                    }
                )

    return occupation_rules


def unique_skill_ids(rules: list[dict[str, Any]]) -> list[str]:
    seen: set[str] = set()
    ids: list[str] = []
    for rule in rules:
        skill_id = rule["skillId"]
        if skill_id in seen:
            continue
        seen.add(skill_id)
        ids.append(skill_id)
    return ids


def collect_occupations(workbook: Any) -> list[dict[str, Any]]:
    sheet = workbook[OCCUPATION_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    skill_rules_by_excel_no = collect_skill_matrix(workbook)
    occupations: list[dict[str, Any]] = []

    for row_index, row_values in enumerate(rows[2:], start=3):
        excel_no = row_values[0] if len(row_values) > 0 else None
        name = clean_text(row_values[1] if len(row_values) > 1 else None)
        if not isinstance(excel_no, (int, float)) or name is None:
            continue

        credit_range = clean_text(row_values[3] if len(row_values) > 3 else None)
        skill_points = clean_text(row_values[4] if len(row_values) > 4 else None) or "教育×4"
        skills_text = clean_text(row_values[6] if len(row_values) > 6 else None) or "详见职业技能矩阵。"
        contacts = clean_text(row_values[10] if len(row_values) > 10 else None)
        raw_description = clean_text(row_values[12] if len(row_values) > 12 else None)
        rules = skill_rules_by_excel_no.get(int(excel_no), [])
        category = classify_occupation(name)
        description = make_display_description(name, category["label"], raw_description, skills_text)

        occupations.append(
            {
                "id": int(excel_no),
                "excelNo": int(excel_no),
                "name": name,
                "aliases": [],
                "creditRange": credit_range,
                "credit": parse_credit_range(credit_range),
                "skillPoints": skill_points,
                "pointFormula": parse_point_formula(skill_points),
                "category": category["label"],
                "icon": icon_for_occupation(name, category),
                "shortDesc": make_short_desc(description, skills_text),
                "skillIds": unique_skill_ids(rules),
                "skillRules": rules,
                "skillsText": skills_text,
                "contacts": contacts,
                "description": description,
                "sourceNote": raw_description if is_source_note(raw_description) else None,
                "source": {"workbook": WORKBOOK_PATH.name, "sheet": OCCUPATION_SHEET, "row": row_index},
            }
        )

    return occupations


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    occupations = collect_occupations(workbook)
    payload = {
        "metadata": {
            "sourceWorkbook": WORKBOOK_PATH.name,
            "sourceSheet": OCCUPATION_SHEET,
            "skillMatrixSheet": OCCUPATION_SKILL_SHEET,
            "recordCount": len(occupations),
        },
        "occupations": occupations,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(occupations)} occupations to {OUTPUT_PATH.relative_to(ROOT_DIR)}")


if __name__ == "__main__":
    main()
